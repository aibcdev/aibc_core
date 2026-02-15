#!/usr/bin/env python3
"""
Cookie.fun Playwright Scraper - Bypasses Cloudflare

Scrapes user data from Cookie.fun using Playwright browser automation.
Extracts: username, user ID, onchain address, score/points.

Usage:
    python cookie3_playwright_scraper.py
"""

import json
import time
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout


class CookieFunPlaywrightScraper:
    """Playwright-based scraper for Cookie.fun."""
    
    BASE_URL = "https://www.cookie.fun"
    
    def __init__(self, headless: bool = False):
        self.headless = headless
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None
        self.all_data = []
        self.api_responses = []
    
    def start_browser(self):
        """Start Playwright browser with stealth settings."""
        print("🌐 Starting Playwright browser...")
        self.playwright = sync_playwright().start()
        
        # Use chromium with stealth-like settings
        self.browser = self.playwright.chromium.launch(
            headless=self.headless,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
            ]
        )
        
        # Create context with realistic viewport and user agent
        self.context = self.browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36"
        )
        
        # Intercept API responses
        self.context.on("response", self._handle_response)
        
        self.page = self.context.new_page()
        print("✅ Browser started")
    
    def _handle_response(self, response):
        """Capture API responses."""
        url = response.url
        if "api.cookie.fun" in url or "cookie" in url.lower():
            try:
                if "application/json" in response.headers.get("content-type", ""):
                    data = response.json()
                    self.api_responses.append({
                        "url": url,
                        "data": data
                    })
                    print(f"  📡 Captured API response from: {url[:80]}...")
            except:
                pass
    
    def wait_for_cloudflare(self, timeout: int = 60):
        """Wait for Cloudflare challenge to complete."""
        print("⏳ Waiting for Cloudflare verification...")
        start = time.time()
        
        while time.time() - start < timeout:
            title = self.page.title()
            content = self.page.content()[:2000]
            
            if "Just a moment" not in title and "Cloudflare" not in content:
                print("✅ Cloudflare bypass successful!")
                time.sleep(2)
                return True
            
            time.sleep(1)
        
        print("❌ Cloudflare bypass timed out")
        return False
    
    def navigate_to_page(self, path: str = "/agents"):
        """Navigate to a specific page."""
        url = f"{self.BASE_URL}{path}"
        print(f"\n📊 Navigating to {url}...")
        
        try:
            # Use 'commit' instead of 'networkidle' to not block on Cloudflare
            self.page.goto(url, wait_until="commit", timeout=30000)
        except Exception as e:
            print(f"  Initial navigation: {e}")
        
        # Wait for Cloudflare with extended timeout
        if self.wait_for_cloudflare(timeout=90):
            # Additional wait for content to load after Cloudflare
            try:
                self.page.wait_for_load_state("domcontentloaded", timeout=30000)
            except:
                pass
            time.sleep(3)
            return True
        return False
    
    def scroll_and_load(self, max_scrolls: int = 30):
        """Scroll to load lazy content."""
        print("📜 Scrolling to load all data...")
        
        for i in range(max_scrolls):
            # Scroll down
            self.page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(1)
            
            # Check if we've reached the end
            scroll_height = self.page.evaluate("document.body.scrollHeight")
            viewport_height = self.page.evaluate("window.innerHeight")
            scroll_top = self.page.evaluate("window.scrollY")
            
            if scroll_top + viewport_height >= scroll_height - 100:
                # Try clicking "Load More" if exists
                try:
                    load_more = self.page.locator("text=Load More, text=Show More, button:has-text('More')")
                    if load_more.count() > 0:
                        load_more.first.click()
                        time.sleep(2)
                        continue
                except:
                    pass
                
                print(f"  Reached end after {i+1} scrolls")
                break
            
            if (i + 1) % 10 == 0:
                print(f"  Scrolled {i+1} times...")
    
    def extract_from_api_responses(self):
        """Extract data from captured API responses."""
        print("\n🔍 Extracting from captured API responses...")
        data = []
        
        for resp in self.api_responses:
            try:
                resp_data = resp.get("data", {})
                
                # Handle different response structures
                items = []
                if isinstance(resp_data, dict):
                    # Look for common patterns
                    for key in ["ok", "data", "agents", "items", "results", "users", "creators"]:
                        if key in resp_data:
                            val = resp_data[key]
                            if isinstance(val, list):
                                items = val
                                break
                            elif isinstance(val, dict) and "data" in val:
                                items = val["data"] if isinstance(val["data"], list) else [val["data"]]
                                break
                elif isinstance(resp_data, list):
                    items = resp_data
                
                for item in items:
                    if isinstance(item, dict):
                        entry = self._parse_item(item)
                        if entry.get("username") or entry.get("onchain_address"):
                            data.append(entry)
                            
            except Exception as e:
                print(f"  Error parsing response: {e}")
        
        print(f"  ✓ Extracted {len(data)} items from API responses")
        return data
    
    def _parse_item(self, item: dict) -> dict:
        """Parse an individual item from API response."""
        # Extract contract address from various possible locations
        address = ""
        if "contracts" in item and isinstance(item["contracts"], list) and item["contracts"]:
            address = item["contracts"][0].get("contractAddress", "")
        if not address:
            address = item.get("walletAddress", item.get("address", item.get("contractAddress", "")))
        
        return {
            "username": item.get("twitterUsername", item.get("twitter", item.get("username", item.get("name", "")))),
            "user_id": item.get("agentId", item.get("id", item.get("userId", ""))),
            "display_name": item.get("name", item.get("displayName", "")),
            "onchain_address": address,
            "mindshare_score": item.get("mindshare", item.get("score", 0)),
            "followers": item.get("followersCount", item.get("followers", 0)),
            "market_cap": item.get("marketCap", 0),
            "volume_24h": item.get("volume24Hours", item.get("volume24h", 0)),
            "price_usd": item.get("priceUsd", item.get("price", 0)),
            "chain": item.get("chain", ""),
            "category": item.get("category", "agent"),
            "source": "cookie.fun"
        }
    
    def extract_from_nextjs_data(self):
        """Extract from Next.js embedded data."""
        print("\n🔍 Checking for Next.js embedded data...")
        data = []
        
        try:
            # Try to get __NEXT_DATA__
            script_content = self.page.evaluate("""
                () => {
                    const script = document.getElementById('__NEXT_DATA__');
                    return script ? script.textContent : null;
                }
            """)
            
            if script_content:
                next_data = json.loads(script_content)
                props = next_data.get("props", {}).get("pageProps", {})
                
                for key in ["agents", "data", "items", "initialData"]:
                    if key in props and isinstance(props[key], list):
                        for item in props[key]:
                            entry = self._parse_item(item)
                            if entry.get("username") or entry.get("onchain_address"):
                                data.append(entry)
                
                print(f"  ✓ Extracted {len(data)} items from Next.js data")
        except Exception as e:
            print(f"  No Next.js data found: {e}")
        
        return data
    
    def extract_from_page_content(self):
        """Extract data from visible page elements."""
        print("\n🔍 Extracting from visible page content...")
        data = []
        
        try:
            # Get all text content
            content = self.page.content()
            
            # Find wallet addresses
            addresses = set(re.findall(r'0x[a-fA-F0-9]{40}', content))
            print(f"  Found {len(addresses)} unique addresses")
            
            # Find Twitter handles
            handles = set(re.findall(r'@[a-zA-Z0-9_]{1,15}', content))
            print(f"  Found {len(handles)} unique handles")
            
            # Try to find table rows or card elements
            rows = self.page.locator("tr, [class*='row'], [class*='item'], [class*='card']").all()
            print(f"  Found {len(rows)} row/card elements")
            
            for row in rows[:200]:  # Limit for performance
                try:
                    text = row.inner_text()
                    if len(text) < 10:
                        continue
                    
                    entry = {
                        "raw_text": text[:300],
                        "source": "cookie.fun"
                    }
                    
                    # Extract data from text
                    row_addresses = re.findall(r'0x[a-fA-F0-9]{40}', text)
                    if row_addresses:
                        entry["onchain_address"] = row_addresses[0]
                    
                    row_handles = re.findall(r'@[a-zA-Z0-9_]{1,15}', text)
                    if row_handles:
                        entry["username"] = row_handles[0]
                    
                    # Extract numbers (potential scores/stats)
                    numbers = re.findall(r'[\d,]+\.?\d*[KMB]?%?', text)
                    if len(numbers) >= 1:
                        entry["stats"] = numbers[:5]
                    
                    if entry.get("username") or entry.get("onchain_address"):
                        data.append(entry)
                except:
                    continue
            
            print(f"  ✓ Extracted {len(data)} items from page content")
        except Exception as e:
            print(f"  Error extracting from content: {e}")
        
        return data
    
    def take_screenshot(self, name: str = "page"):
        """Save screenshot."""
        filename = f"cookie_screenshot_{name}_{int(time.time())}.png"
        self.page.screenshot(path=filename, full_page=True)
        print(f"📸 Screenshot saved: {filename}")
        return filename
    
    def close(self):
        """Close browser."""
        if self.context:
            self.context.close()
        if self.browser:
            self.browser.close()
        if self.playwright:
            self.playwright.stop()
        print("🔒 Browser closed")


def export_to_excel(data: list, filename: str = "cookie3_users.xlsx"):
    """Export to Excel with formatting."""
    if not data:
        print("No data to export!")
        return
    
    # Convert to DataFrame and clean up
    df = pd.DataFrame(data)
    
    # Remove duplicates
    if "username" in df.columns:
        df = df.drop_duplicates(subset=["username"], keep="first")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cookie3 Users"
    
    # Style headers
    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Write data
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(value, (list, dict)):
                cell.value = str(value)
            else:
                cell.value = value if value else ""
    
    # Auto column width
    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 25
    
    ws.freeze_panes = "A2"
    wb.save(filename)
    print(f"\n✅ Exported {len(df)} unique records to {filename}")


def main():
    print("=" * 60)
    print("Cookie.fun Playwright Scraper")
    print("=" * 60)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    scraper = CookieFunPlaywrightScraper(headless=False)
    all_data = []
    
    try:
        scraper.start_browser()
        
        # Navigate to agents page
        if scraper.navigate_to_page("/agents"):
            scraper.take_screenshot("agents_initial")
            
            # Scroll to load content
            scraper.scroll_and_load(max_scrolls=20)
            scraper.take_screenshot("agents_scrolled")
            
            # Extract data from multiple sources
            api_data = scraper.extract_from_api_responses()
            all_data.extend(api_data)
            
            nextjs_data = scraper.extract_from_nextjs_data()
            all_data.extend(nextjs_data)
            
            page_data = scraper.extract_from_page_content()
            all_data.extend(page_data)
        else:
            print("❌ Failed to load page")
            scraper.take_screenshot("failed")
    
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        scraper.close()
    
    # Export
    print("\n" + "=" * 60)
    print("EXPORT")
    print("=" * 60)
    
    if all_data:
        export_to_excel(all_data, "cookie3_users.xlsx")
        print(f"\n📈 Summary: {len(all_data)} total items collected")
    else:
        print("\n⚠️ No data collected. Check screenshots for page state.")
    
    print(f"\nCompleted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
