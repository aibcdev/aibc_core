#!/usr/bin/env python3
"""
COOKIE Token Holder Scraper via Blockscout API

Scrapes COOKIE token holder data from Base Blockscout - NO API KEY REQUIRED!
Token Address: 0xc0041ef357b183448b235a8ea73ce4e4ec8c265f (on Base chain)

Usage:
    python cookie_blockscout_scraper.py
"""

import json
import time
from datetime import datetime
from typing import List, Dict, Optional

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# COOKIE Token on Base
COOKIE_TOKEN_ADDRESS = "0xc0041ef357b183448b235a8ea73ce4e4ec8c265f"
BLOCKSCOUT_API = "https://base.blockscout.com/api/v2"


class CookieBlockscoutScraper:
    """Scraper for COOKIE token data via Blockscout API."""
    
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            "Accept": "application/json"
        })
    
    def get_token_info(self) -> Dict:
        """Get token metadata."""
        print("📊 Fetching COOKIE token info...")
        url = f"{BLOCKSCOUT_API}/tokens/{COOKIE_TOKEN_ADDRESS}"
        
        try:
            response = self.session.get(url, timeout=30)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            print(f"  Error: {e}")
            return {}
    
    def get_holders(self, next_page_params: Optional[str] = None) -> Dict:
        """Get token holders page."""
        url = f"{BLOCKSCOUT_API}/tokens/{COOKIE_TOKEN_ADDRESS}/holders"
        params = {}
        
        if next_page_params:
            # Parse the next page params
            params = next_page_params
        
        try:
            response = self.session.get(url, params=params, timeout=30)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            print(f"  Error: {e}")
            return {}
    
    def get_all_holders(self, max_pages: int = 100) -> List[Dict]:
        """Fetch all token holders with pagination."""
        print("\n📋 Fetching all COOKIE token holders...")
        
        all_holders = []
        next_params = None
        page = 1
        
        while page <= max_pages:
            print(f"  Fetching page {page}...", end=" ")
            
            data = self.get_holders(next_params)
            
            if not data or "items" not in data:
                print("No more data")
                break
            
            items = data.get("items", [])
            if not items:
                print("Empty page")
                break
            
            for item in items:
                holder = self._parse_holder(item)
                if holder:
                    all_holders.append(holder)
            
            print(f"Got {len(items)} holders (total: {len(all_holders)})")
            
            # Check for next page
            next_page = data.get("next_page_params")
            if not next_page:
                print("  Reached last page")
                break
            
            next_params = next_page
            page += 1
            time.sleep(0.3)  # Rate limiting
        
        return all_holders
    
    def _parse_holder(self, item: Dict) -> Optional[Dict]:
        """Parse a holder item from API response."""
        try:
            address_info = item.get("address", {})
            value_raw = item.get("value", "0")
            
            # Convert to decimal (18 decimals)
            balance = int(value_raw) / (10 ** 18)
            
            # Extract metadata/tags
            metadata = address_info.get("metadata", {}) or {}
            tags = metadata.get("tags", [])
            tag_names = [t.get("name", "") for t in tags if t.get("name")]
            
            return {
                "onchain_address": address_info.get("hash", ""),
                "balance": balance,
                "balance_raw": value_raw,
                "is_contract": address_info.get("is_contract", False),
                "is_verified": address_info.get("is_verified", False),
                "contract_name": address_info.get("name", ""),
                "tags": ", ".join(tag_names) if tag_names else "",
                "ens_domain": address_info.get("ens_domain_name", ""),
                "source": "blockscout.com"
            }
        except Exception as e:
            return None
    
    def get_token_transfers(self, next_page_params: Optional[Dict] = None) -> Dict:
        """Get token transfers for additional context."""
        url = f"{BLOCKSCOUT_API}/tokens/{COOKIE_TOKEN_ADDRESS}/transfers"
        params = next_page_params or {}
        
        try:
            response = self.session.get(url, params=params, timeout=30)
            response.raise_for_status()
            return response.json()
        except Exception as e:
            print(f"  Error: {e}")
            return {}


def export_to_excel(data: list, token_info: dict, filename: str = "cookie3_users.xlsx"):
    """Export to formatted Excel with summary sheet."""
    if not data:
        print("No data to export!")
        return
    
    df = pd.DataFrame(data)
    
    # Sort by balance
    df = df.sort_values("balance", ascending=False)
    
    # Create workbook
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    ws_summary["A1"] = "COOKIE Token Holder Report"
    ws_summary["A1"].font = Font(size=16, bold=True)
    
    ws_summary["A3"] = "Token Name:"
    ws_summary["B3"] = token_info.get("name", "COOKIE")
    ws_summary["A4"] = "Symbol:"
    ws_summary["B4"] = token_info.get("symbol", "COOKIE")
    ws_summary["A5"] = "Contract:"
    ws_summary["B5"] = COOKIE_TOKEN_ADDRESS
    ws_summary["A6"] = "Chain:"
    ws_summary["B6"] = "Base"
    ws_summary["A7"] = "Total Holders:"
    ws_summary["B7"] = len(df)
    ws_summary["A8"] = "Report Date:"
    ws_summary["B8"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Top 10 holders
    ws_summary["A10"] = "Top 10 Holders"
    ws_summary["A10"].font = Font(bold=True)
    
    for i, (_, row) in enumerate(df.head(10).iterrows(), start=11):
        ws_summary[f"A{i}"] = f"{i-10}. {row['onchain_address'][:20]}..."
        ws_summary[f"B{i}"] = f"{row['balance']:,.2f} COOKIE"
        if row.get("tags"):
            ws_summary[f"C{i}"] = row["tags"]
    
    # Holders sheet
    ws_holders = wb.create_sheet("All Holders")
    
    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Column order
    columns = ["onchain_address", "balance", "is_contract", "contract_name", "tags", "ens_domain"]
    display_names = ["Wallet Address", "COOKIE Balance", "Is Contract", "Contract Name", "Tags", "ENS Domain"]
    
    for col_idx, name in enumerate(display_names, 1):
        cell = ws_holders.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Write data
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col in enumerate(columns, 1):
            value = row.get(col, "")
            cell = ws_holders.cell(row=row_idx, column=col_idx)
            
            if col == "balance":
                cell.value = round(value, 4) if value else 0
                cell.number_format = "#,##0.0000"
            elif isinstance(value, bool):
                cell.value = "Yes" if value else "No"
            else:
                cell.value = str(value) if value else ""
    
    # Column widths
    ws_holders.column_dimensions["A"].width = 45
    ws_holders.column_dimensions["B"].width = 20
    ws_holders.column_dimensions["C"].width = 12
    ws_holders.column_dimensions["D"].width = 30
    ws_holders.column_dimensions["E"].width = 25
    ws_holders.column_dimensions["F"].width = 25
    
    ws_holders.freeze_panes = "A2"
    
    wb.save(filename)
    print(f"\n✅ Exported {len(df)} holders to {filename}")


def main():
    print("=" * 60)
    print("COOKIE Token Holder Scraper (Blockscout API)")
    print("=" * 60)
    print(f"Token: {COOKIE_TOKEN_ADDRESS}")
    print(f"Chain: Base")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    scraper = CookieBlockscoutScraper()
    
    # Get token info
    token_info = scraper.get_token_info()
    if token_info:
        print(f"  Token: {token_info.get('name', 'N/A')} ({token_info.get('symbol', 'N/A')})")
        print(f"  Total Supply: {int(token_info.get('total_supply', 0)) / 10**18:,.0f}")
        print(f"  Holders Count: {token_info.get('holders', 'N/A')}")
    
    # Get all holders
    holders = scraper.get_all_holders(max_pages=100)
    
    # Export
    print("\n" + "=" * 60)
    print("EXPORT")
    print("=" * 60)
    
    if holders:
        export_to_excel(holders, token_info, "cookie3_users.xlsx")
        
        # Stats
        df = pd.DataFrame(holders)
        contracts = df[df["is_contract"] == True]
        wallets = df[df["is_contract"] == False]
        
        print(f"\n📈 Summary:")
        print(f"   Total holders: {len(holders)}")
        print(f"   Wallets: {len(wallets)}")
        print(f"   Contracts: {len(contracts)}")
        print(f"   Top holder: {holders[0]['balance']:,.2f} COOKIE")
        
        if holders[0].get("tags"):
            print(f"   Top holder tags: {holders[0]['tags']}")
    else:
        print("\n❌ No holder data collected.")
    
    print(f"\nCompleted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
