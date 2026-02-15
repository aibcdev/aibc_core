#!/usr/bin/env python3
"""
Cookie.fun Selenium Scraper - Bypasses Cloudflare

Scrapes user data from Cookie.fun using browser automation to bypass Cloudflare protection.
Extracts: username, user ID, onchain address, score/points.

Usage:
    python cookie3_selenium_scraper.py
"""

import json
import time
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    import undetected_chromedriver as uc
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException, NoSuchElementException
except ImportError:
    print("Installing required packages...")
    import subprocess
    subprocess.run(["pip", "install", "undetected-chromedriver", "selenium", "-q"])
    import undetected_chromedriver as uc
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import TimeoutException, NoSuchElementException


class CookieFunSeleniumScraper:
    """Selenium-based scraper for Cookie.fun with Cloudflare bypass."""
    
    BASE_URL = "https://www.cookie.fun"
    
    def __init__(self, headless: bool = False):
        """Initialize undetected Chrome browser."""
        self.headless = headless
        self.driver = None
        self.all_data = []
    
    def start_browser(self):
        """Start undetected Chrome browser."""
        print("🌐 Starting browser (bypassing Cloudflare)...")
        options = uc.ChromeOptions()
        if self.headless:
            options.add_argument("--headless")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--window-size=1920,1080")
        
        self.driver = uc.Chrome(options=options)
        self.driver.implicitly_wait(10)
    
    def wait_for_cloudflare(self, timeout: int = 30):
        """Wait for Cloudflare challenge to complete."""
        print("⏳ Waiting for Cloudflare verification...")
        try:
            # Wait until we're past Cloudflare (check for main content)
            WebDriverWait(self.driver, timeout).until(
                lambda d: "Just a moment" not in d.title and "Cloudflare" not in d.page_source[:1000]
            )
            print("✅ Cloudflare bypass successful!")
            time.sleep(2)  # Extra wait for page to fully load
            return True
        except TimeoutException:
            print("❌ Cloudflare bypass timed out")
            return False
    
    def navigate_to_agents(self):
        """Navigate to the agents page."""
        print(f"\n📊 Navigating to {self.BASE_URL}/agents...")
        self.driver.get(f"{self.BASE_URL}/agents")
        return self.wait_for_cloudflare()
    
    def scroll_and_load_all(self, max_scrolls: int = 50):
        """Scroll to load all lazy-loaded content."""
        print("📜 Scrolling to load all data...")
        last_height = self.driver.execute_script("return document.body.scrollHeight")
        scroll_count = 0
        
        while scroll_count < max_scrolls:
            # Scroll down
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1.5)
            
            # Check for new content
            new_height = self.driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                print(f"  Reached end after {scroll_count} scrolls")
                break
            last_height = new_height
            scroll_count += 1
            
            if scroll_count % 10 == 0:
                print(f"  Scrolled {scroll_count} times...")
    
    def extract_agent_data(self):
        """Extract agent data from the current page."""
        print("🔍 Extracting agent data from page...")
        agents = []
        
        # Try to capture network requests for API data
        try:
            # Look for agent cards/rows in the page
            # These selectors may need adjustment based on actual page structure
            possible_selectors = [
                "[data-testid='agent-row']",
                ".agent-card",
                ".agent-row",
                "tr[class*='agent']",
                "[class*='AgentCard']",
                "[class*='agent-item']",
                "table tbody tr",
                ".leaderboard-row",
                "[class*='row']"
            ]
            
            for selector in possible_selectors:
                try:
                    elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    if elements:
                        print(f"  Found {len(elements)} elements with selector: {selector}")
                        for elem in elements[:5]:  # Sample first 5
                            print(f"    Text: {elem.text[:100]}...")
                except:
                    continue
            
            # Get page source for debugging
            page_source = self.driver.page_source
            
            # Try to find JSON data embedded in the page
            if '"agents"' in page_source or '"data"' in page_source:
                print("  Found potential JSON data in page source")
                # Look for Next.js data
                try:
                    scripts = self.driver.find_elements(By.CSS_SELECTOR, "script[id='__NEXT_DATA__']")
                    for script in scripts:
                        data = json.loads(script.get_attribute('innerHTML'))
                        print(f"  Found Next.js data: {list(data.get('props', {}).keys())}")
                        agents.extend(self._parse_nextjs_data(data))
                except Exception as e:
                    print(f"  Error parsing Next.js data: {e}")
            
            # Also try to intercept XHR responses by checking for API patterns
            try:
                # Execute JavaScript to get any fetched data
                result = self.driver.execute_script("""
                    if (window.__NEXT_DATA__) {
                        return JSON.stringify(window.__NEXT_DATA__);
                    }
                    return null;
                """)
                if result:
                    data = json.loads(result)
                    agents.extend(self._parse_nextjs_data(data))
            except Exception as e:
                pass
                
        except Exception as e:
            print(f"  Error extracting data: {e}")
        
        return agents
    
    def _parse_nextjs_data(self, data: dict) -> list:
        """Parse Next.js embedded data structure."""
        agents = []
        try:
            # Navigate through common Next.js data structures
            props = data.get('props', {})
            page_props = props.get('pageProps', {})
            
            # Look for agent data in various possible locations
            possible_keys = ['agents', 'data', 'items', 'results', 'leaderboard']
            for key in possible_keys:
                if key in page_props:
                    items = page_props[key]
                    if isinstance(items, list):
                        for item in items:
                            agents.append({
                                'username': item.get('twitterUsername', item.get('name', item.get('username', ''))),
                                'user_id': item.get('id', item.get('agentId', '')),
                                'display_name': item.get('name', item.get('displayName', '')),
                                'onchain_address': self._extract_address(item),
                                'mindshare_score': item.get('mindshare', item.get('score', 0)),
                                'followers': item.get('followersCount', item.get('followers', 0)),
                                'market_cap': item.get('marketCap', 0),
                                'category': 'agent',
                                'source': 'cookie.fun/agents'
                            })
        except Exception as e:
            print(f"  Error parsing Next.js data: {e}")
        return agents
    
    def _extract_address(self, item: dict) -> str:
        """Extract wallet/contract address from item."""
        if 'contracts' in item and item['contracts']:
            return item['contracts'][0].get('contractAddress', '')
        return item.get('walletAddress', item.get('address', item.get('contractAddress', '')))
    
    def scrape_visible_table_data(self):
        """Scrape any visible table data on the page."""
        print("📋 Scraping visible table data...")
        data = []
        
        try:
            # Wait for page content to load
            time.sleep(3)
            
            # Try to find any tables
            tables = self.driver.find_elements(By.TAG_NAME, "table")
            print(f"  Found {len(tables)} tables")
            
            for table_idx, table in enumerate(tables):
                try:
                    rows = table.find_elements(By.TAG_NAME, "tr")
                    print(f"  Table {table_idx}: {len(rows)} rows")
                    
                    for row in rows[1:]:  # Skip header
                        cells = row.find_elements(By.TAG_NAME, "td")
                        if cells:
                            row_data = {
                                'raw_text': row.text,
                                'cells': [c.text for c in cells],
                                'source': 'cookie.fun'
                            }
                            # Try to extract structured data
                            text = row.text
                            # Look for wallet addresses (0x...)
                            import re
                            addresses = re.findall(r'0x[a-fA-F0-9]{40}', text)
                            if addresses:
                                row_data['onchain_address'] = addresses[0]
                            # Look for Twitter handles (@...)
                            handles = re.findall(r'@[\w]+', text)
                            if handles:
                                row_data['username'] = handles[0]
                            data.append(row_data)
                except Exception as e:
                    print(f"  Error parsing table: {e}")
            
            # Also try to find list/card items
            cards = self.driver.find_elements(By.CSS_SELECTOR, "[class*='card'], [class*='item'], [class*='agent']")
            print(f"  Found {len(cards)} card/item elements")
            
            for card in cards[:100]:  # Limit for performance
                try:
                    text = card.text.strip()
                    if text and len(text) > 10:
                        import re
                        entry = {
                            'raw_text': text[:500],
                            'source': 'cookie.fun'
                        }
                        addresses = re.findall(r'0x[a-fA-F0-9]{40}', text)
                        if addresses:
                            entry['onchain_address'] = addresses[0]
                        handles = re.findall(r'@[\w]+', text)
                        if handles:
                            entry['username'] = handles[0]
                        # Look for numbers that might be scores
                        numbers = re.findall(r'\d+\.?\d*[KMB]?', text)
                        if numbers:
                            entry['score'] = numbers[0] if numbers else ''
                        data.append(entry)
                except:
                    continue
                    
        except Exception as e:
            print(f"  Error scraping table data: {e}")
        
        return data
    
    def take_screenshot(self, name: str = "page"):
        """Take screenshot for debugging."""
        filename = f"cookie_screenshot_{name}_{int(time.time())}.png"
        self.driver.save_screenshot(filename)
        print(f"📸 Screenshot saved: {filename}")
        return filename
    
    def close(self):
        """Close the browser."""
        if self.driver:
            self.driver.quit()
            print("🔒 Browser closed")


def export_to_excel(data: list, filename: str = "cookie3_users.xlsx"):
    """Export data to formatted Excel file."""
    if not data:
        print("No data to export!")
        return
    
    df = pd.DataFrame(data)
    
    # Create workbook with styling
    wb = Workbook()
    ws = wb.active
    ws.title = "Cookie3 Users"
    
    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Write headers
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Write data
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
    
    # Auto-width columns
    for col_idx, column in enumerate(df.columns, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20
    
    ws.freeze_panes = "A2"
    wb.save(filename)
    print(f"\n✅ Exported {len(data)} records to {filename}")


def main():
    print("=" * 60)
    print("Cookie.fun Selenium Scraper (Cloudflare Bypass)")
    print("=" * 60)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    scraper = CookieFunSeleniumScraper(headless=False)
    all_data = []
    
    try:
        scraper.start_browser()
        
        # Navigate to agents page
        if scraper.navigate_to_agents():
            # Take screenshot to see what we have
            scraper.take_screenshot("agents_page")
            
            # Scroll to load all content
            scraper.scroll_and_load_all(max_scrolls=20)
            
            # Extract data
            agents = scraper.extract_agent_data()
            if agents:
                all_data.extend(agents)
                print(f"  ✓ Extracted {len(agents)} agents from embedded data")
            
            # Also scrape visible table/card data
            visible_data = scraper.scrape_visible_table_data()
            if visible_data:
                all_data.extend(visible_data)
                print(f"  ✓ Scraped {len(visible_data)} items from visible elements")
            
            # Take final screenshot
            scraper.take_screenshot("final")
        else:
            print("❌ Failed to bypass Cloudflare")
            scraper.take_screenshot("cloudflare_blocked")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        scraper.close()
    
    # Export results
    print("\n" + "=" * 60)
    print("EXPORT")
    print("=" * 60)
    
    if all_data:
        # Remove duplicates
        df = pd.DataFrame(all_data)
        if 'username' in df.columns:
            df = df.drop_duplicates(subset=['username'], keep='first')
        unique_data = df.to_dict('records')
        
        export_to_excel(unique_data, "cookie3_users.xlsx")
        
        print(f"\n📈 Summary:")
        print(f"   Total records: {len(unique_data)}")
    else:
        print("\n⚠️ No data collected. The page structure may have changed.")
        print("Check the screenshots to see what's visible on the page.")
    
    print(f"\nCompleted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
