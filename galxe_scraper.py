#!/usr/bin/env python3
"""Quick Galxe Leaderboard Scraper - NO API KEY NEEDED"""
import requests
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

url = 'https://graphigo.prd.galaxy.eco/query'

# Get popular spaces
spaces_query = '{ spaces(first: 50) { list { id name alias } } }'
resp = requests.post(url, json={'query': spaces_query}, timeout=30)
result = resp.json()
if 'errors' in result:
    print(f"API Error: {result['errors']}")
    exit(1)
spaces = result.get('data', {}).get('spaces', {}).get('list', [])

print(f"Found {len(spaces)} spaces")

all_users = []

for space in spaces[:20]:  # Top 20 spaces
    space_id = int(space['id'])
    print(f"Scraping {space['name']} (ID: {space_id})...")
    
    lb_query = f'''
    {{
      space(id: {space_id}) {{
        name
        loyaltyPointsRanks(first: 100) {{
          list {{
            address
            rank
            points
          }}
        }}
      }}
    }}
    '''
    
    try:
        lb_resp = requests.post(url, json={'query': lb_query}, timeout=30)
        data = lb_resp.json()
        
        if 'data' in data and data['data']['space']:
            ranks = data['data']['space'].get('loyaltyPointsRanks', {}).get('list', [])
            for r in ranks:
                all_users.append({
                    'space': space['name'],
                    'onchain_address': r['address'],
                    'rank': r['rank'],
                    'points': r['points'],
                    'source': 'galxe.com'
                })
            print(f"  Got {len(ranks)} users")
    except:
        pass

print(f"\nTotal users: {len(all_users)}")

# Export
df = pd.DataFrame(all_users)
df = df.drop_duplicates(subset=['onchain_address'], keep='first')

wb = Workbook()
ws = wb.active
ws.title = "Galxe Users"

header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col_idx, column in enumerate(df.columns, 1):
    cell = ws.cell(row=1, column=col_idx, value=column.replace("_", " ").title())
    cell.fill = header_fill
    cell.font = header_font

for row_idx, row in enumerate(df.values, 2):
    for col_idx, value in enumerate(row, 1):
        ws.cell(row=row_idx, column=col_idx, value=str(value))

for col in ws.columns:
    ws.column_dimensions[col[0].column_letter].width = 30

ws.freeze_panes = "A2"
wb.save("galxe_users.xlsx")
print(f"✅ Exported {len(df)} unique users to galxe_users.xlsx")
