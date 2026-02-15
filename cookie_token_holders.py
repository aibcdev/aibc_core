#!/usr/bin/env python3
"""
COOKIE Token Holder Scraper via BaseScan

Scrapes COOKIE token holder data from BaseScan blockchain explorer.
Token Address: 0xc0041ef357b183448b235a8ea73ce4e4ec8c265f (on Base chain)

Usage:
    python cookie_token_holders.py
"""

import json
import time
import re
from datetime import datetime
from typing import List, Dict

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# COOKIE Token on Base
COOKIE_TOKEN_ADDRESS = "0xc0041ef357b183448b235a8ea73ce4e4ec8c265f"

# BaseScan API (free tier)
BASESCAN_API_URL = "https://api.basescan.org/api"

# Alternative: Direct page scraping endpoints
BASESCAN_HOLDERS_URL = f"https://basescan.org/token/{COOKIE_TOKEN_ADDRESS}#balances"


class CookieTokenScraper:
    """Scraper for COOKIE token holder data."""
    
    def __init__(self, api_key: str = None):
        self.api_key = api_key
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
        })
    
    def get_token_info(self) -> Dict:
        """Get basic token info from BaseScan API."""
        print("📊 Fetching COOKIE token info...")
        
        params = {
            "module": "token",
            "action": "tokeninfo",
            "contractaddress": COOKIE_TOKEN_ADDRESS,
        }
        if self.api_key:
            params["apikey"] = self.api_key
        
        try:
            response = self.session.get(BASESCAN_API_URL, params=params, timeout=30)
            data = response.json()
            if data.get("status") == "1":
                result = data.get("result", [{}])
                if isinstance(result, list) and result:
                    return result[0]
                return result
        except Exception as e:
            print(f"  Error: {e}")
        
        return {}
    
    def get_top_holders_via_api(self, page: int = 1, offset: int = 100) -> List[Dict]:
        """Get top token holders via BaseScan API."""
        print(f"📋 Fetching top holders (page {page})...")
        
        # Note: Top holders endpoint may require API key
        params = {
            "module": "token",
            "action": "tokenholderlist",
            "contractaddress": COOKIE_TOKEN_ADDRESS,
            "page": page,
            "offset": offset,
        }
        if self.api_key:
            params["apikey"] = self.api_key
        
        try:
            response = self.session.get(BASESCAN_API_URL, params=params, timeout=30)
            data = response.json()
            
            if data.get("status") == "1":
                return data.get("result", [])
            else:
                print(f"  API Error: {data.get('message', 'Unknown error')}")
        except Exception as e:
            print(f"  Request Error: {e}")
        
        return []
    
    def get_token_transfers(self, start_block: int = 0, end_block: int = 99999999, page: int = 1) -> List[Dict]:
        """Get token transfer events to identify unique holders."""
        print(f"📤 Fetching token transfers (page {page})...")
        
        params = {
            "module": "account",
            "action": "tokentx",
            "contractaddress": COOKIE_TOKEN_ADDRESS,
            "startblock": start_block,
            "endblock": end_block,
            "page": page,
            "offset": 1000,
            "sort": "desc",
        }
        if self.api_key:
            params["apikey"] = self.api_key
        
        try:
            response = self.session.get(BASESCAN_API_URL, params=params, timeout=30)
            data = response.json()
            
            if data.get("status") == "1":
                return data.get("result", [])
            else:
                print(f"  API message: {data.get('message', 'Unknown')}")
        except Exception as e:
            print(f"  Error: {e}")
        
        return []
    
    def extract_unique_addresses_from_transfers(self, max_pages: int = 10) -> List[str]:
        """Extract unique addresses from transfer history."""
        print("\n🔍 Extracting unique addresses from transfers...")
        
        all_addresses = set()
        
        for page in range(1, max_pages + 1):
            transfers = self.get_token_transfers(page=page)
            
            if not transfers:
                break
            
            for tx in transfers:
                all_addresses.add(tx.get("from", "").lower())
                all_addresses.add(tx.get("to", "").lower())
            
            print(f"  Page {page}: Found {len(transfers)} transfers, {len(all_addresses)} unique addresses")
            time.sleep(0.3)  # Rate limiting
        
        # Remove empty strings and null address
        all_addresses.discard("")
        all_addresses.discard("0x0000000000000000000000000000000000000000")
        
        return list(all_addresses)
    
    def get_address_balance(self, address: str) -> int:
        """Get COOKIE token balance for an address."""
        params = {
            "module": "account",
            "action": "tokenbalance",
            "contractaddress": COOKIE_TOKEN_ADDRESS,
            "address": address,
            "tag": "latest",
        }
        if self.api_key:
            params["apikey"] = self.api_key
        
        try:
            response = self.session.get(BASESCAN_API_URL, params=params, timeout=10)
            data = response.json()
            
            if data.get("status") == "1":
                return int(data.get("result", 0))
        except:
            pass
        
        return 0
    
    def get_holder_data(self, addresses: List[str], max_addresses: int = 500) -> List[Dict]:
        """Get balance data for list of addresses."""
        print(f"\n💰 Fetching balances for {min(len(addresses), max_addresses)} addresses...")
        
        holders = []
        
        for i, address in enumerate(addresses[:max_addresses]):
            balance = self.get_address_balance(address)
            
            if balance > 0:
                holders.append({
                    "onchain_address": address,
                    "balance_raw": balance,
                    "balance": balance / (10 ** 18),  # Assuming 18 decimals
                    "source": "basescan.org"
                })
            
            if (i + 1) % 50 == 0:
                print(f"  Processed {i + 1} addresses, found {len(holders)} holders")
            
            time.sleep(0.25)  # Rate limiting for free API
        
        # Sort by balance
        holders.sort(key=lambda x: x["balance"], reverse=True)
        
        return holders


def export_to_excel(data: list, filename: str = "cookie_token_holders.xlsx"):
    """Export to formatted Excel."""
    if not data:
        print("No data to export!")
        return
    
    df = pd.DataFrame(data)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "COOKIE Token Holders"
    
    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(value, float):
                cell.value = round(value, 4)
            else:
                cell.value = str(value) if value else ""
    
    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 45
    
    ws.freeze_panes = "A2"
    wb.save(filename)
    print(f"\n✅ Exported {len(data)} holders to {filename}")


def main():
    print("=" * 60)
    print("COOKIE Token Holder Scraper (Base Chain)")
    print("=" * 60)
    print(f"Token: {COOKIE_TOKEN_ADDRESS}")
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    scraper = CookieTokenScraper()
    
    # Get token info
    token_info = scraper.get_token_info()
    if token_info:
        print(f"  Token Name: {token_info.get('tokenName', 'N/A')}")
        print(f"  Symbol: {token_info.get('symbol', 'N/A')}")
        print(f"  Decimals: {token_info.get('divisor', 'N/A')}")
    
    # Try direct holder list first
    holders = scraper.get_top_holders_via_api()
    
    if not holders:
        print("\n⚠️ Direct holder list not available without API key")
        print("   Extracting addresses from transfer history instead...")
        
        # Get unique addresses from transfers
        addresses = scraper.extract_unique_addresses_from_transfers(max_pages=10)
        print(f"\n📊 Found {len(addresses)} unique addresses in transfers")
        
        # Get balances (limited due to rate limiting)
        holders = scraper.get_holder_data(addresses, max_addresses=200)
    else:
        # Format API response
        holders = [{
            "onchain_address": h.get("TokenHolderAddress", ""),
            "balance": float(h.get("TokenHolderQuantity", 0)) / (10 ** 18),
            "source": "basescan.org"
        } for h in holders]
    
    # Export
    print("\n" + "=" * 60)
    print("EXPORT")
    print("=" * 60)
    
    if holders:
        export_to_excel(holders, "cookie_token_holders.xlsx")
        
        print(f"\n📈 Summary:")
        print(f"   Total holders found: {len(holders)}")
        print(f"   Top holder balance: {holders[0]['balance']:,.2f} COOKIE" if holders else "N/A")
    else:
        print("\n❌ No holder data collected.")
        print("   Consider using a BaseScan API key for better access.")
        print("   Get one free at: https://basescan.org/apis")
    
    print(f"\nCompleted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
