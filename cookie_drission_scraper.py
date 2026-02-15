#!/usr/bin/env python3
"""
Cookie.fun Scraper using DrissionPage for Cloudflare bypass

Extracts: username, user ID, onchain address, score/points
"""

import json
import time
import re
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from DrissionPage import ChromiumPage
from DrissionPage import ChromiumOptions


def create_browser():
    """Create stealth browser."""
    co = ChromiumOptions()
    co.set_argument('--no-sandbox')
    co.set_argument('--disable-gpu')
    co.set_argument('--window-size=1920,1080')
    # Make it look like a real browser
    co.set_user_agent('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    
    page = ChromiumPage(co)
    return page


def wait_for_cloudflare(page, timeout=60):
    """Wait for Cloudflare to pass."""
    print("⏳ Waiting for Cloudflare...")
    start = time.time()
    
    while time.time() - start < timeout:
        title = page.title
        html = page.html[:2000] if page.html else ""
        
        if "Just a moment" not in title and "Cloudflare" not in html:
            print("✅ Cloudflare bypassed!")
            return True
        
        time.sleep(2)
    
    print("❌ Cloudflare timeout")
    return False


def extract_agents_data(page):
    """Extract agent data from the page."""
    print("🔍 Extracting agent data...")
    agents = []
    
    try:
        # Wait for content to load
        time.sleep(5)
        
        # Try to get __NEXT_DATA__
        try:
            script = page.ele('tag:script@id=__NEXT_DATA__', timeout=5)
            if script:
                data = json.loads(script.text)
                props = data.get('props', {}).get('pageProps', {})
                print(f"  Found Next.js data with keys: {list(props.keys())}")
                
                # Look for agent data
                for key in ['agents', 'data', 'items', 'initialData']:
                    if key in props:
                        items = props[key]
                        if isinstance(items, list):
                            for item in items:
                                agent = parse_agent(item)
                                if agent:
                                    agents.append(agent)
                        elif isinstance(items, dict) and 'data' in items:
                            for item in items['data']:
                                agent = parse_agent(item)
                                if agent:
                                    agents.append(agent)
        except Exception as e:
            print(f"  No Next.js data: {e}")
        
        # Try to intercept network requests
        # This captures API responses the page makes
        try:
            # Look for any visible table/list data
            rows = page.eles('tag:tr') or []
            print(f"  Found {len(rows)} table rows")
            
            for row in rows[1:100]:  # Skip header, limit
                text = row.text
                if text and len(text) > 20:
                    agent = parse_row_text(text)
                    if agent.get('username') or agent.get('onchain_address'):
                        agents.append(agent)
        except Exception as e:
            print(f"  Error parsing rows: {e}")
        
        # Try card elements
        try:
            cards = page.eles('css:[class*=agent], css:[class*=card], css:[class*=item]') or []
            print(f"  Found {len(cards)} card elements")
            
            for card in cards[:200]:
                text = card.text
                if text and len(text) > 20:
                    agent = parse_row_text(text)
                    if agent.get('username') or agent.get('onchain_address'):
                        agents.append(agent)
        except Exception as e:
            print(f"  Error parsing cards: {e}")
        
        # Get full page HTML and extract JSON
        html = page.html
        
        # Look for embedded JSON data
        json_matches = re.findall(r'"agentId":"([^"]+)"', html)
        if json_matches:
            print(f"  Found {len(json_matches)} agent IDs in HTML")
        
        # Try to find agent data in scripts
        scripts = page.eles('tag:script') or []
        for script in scripts:
            try:
                text = script.text or ""
                if '"agents"' in text or '"agentId"' in text:
                    # Try to extract JSON
                    matches = re.findall(r'\{[^{}]*"agentId"[^{}]*\}', text)
                    for match in matches[:100]:
                        try:
                            item = json.loads(match)
                            agent = parse_agent(item)
                            if agent:
                                agents.append(agent)
                        except:
                            continue
            except:
                continue
                
    except Exception as e:
        print(f"  Error: {e}")
    
    return agents


def parse_agent(item):
    """Parse agent item from JSON."""
    if not isinstance(item, dict):
        return None
    
    address = ""
    if 'contracts' in item and item['contracts']:
        if isinstance(item['contracts'], list) and item['contracts']:
            address = item['contracts'][0].get('contractAddress', '')
    if not address:
        address = item.get('walletAddress', item.get('address', item.get('contractAddress', '')))
    
    return {
        'username': item.get('twitterUsername', item.get('twitter', item.get('name', ''))),
        'user_id': item.get('agentId', item.get('id', '')),
        'display_name': item.get('name', item.get('displayName', '')),
        'onchain_address': address,
        'mindshare_score': item.get('mindshare', 0),
        'followers': item.get('followersCount', 0),
        'market_cap': item.get('marketCap', 0),
        'source': 'cookie.fun'
    }


def parse_row_text(text):
    """Parse text from a row/card element."""
    entry = {'source': 'cookie.fun', 'raw_text': text[:200]}
    
    # Extract wallet addresses
    addresses = re.findall(r'0x[a-fA-F0-9]{40}', text)
    if addresses:
        entry['onchain_address'] = addresses[0]
    
    # Extract Twitter handles
    handles = re.findall(r'@([a-zA-Z0-9_]{1,15})', text)
    if handles:
        entry['username'] = handles[0]
    
    # Look for numbers (scores, ranks)
    numbers = re.findall(r'\d+\.?\d*[KMB]?', text)
    if numbers:
        entry['score'] = numbers[0] if numbers else ''
    
    return entry


def export_to_excel(data, filename="cookie3_users.xlsx"):
    """Export to Excel."""
    if not data:
        print("No data to export!")
        return
    
    df = pd.DataFrame(data)
    
    # Remove duplicates
    if 'username' in df.columns:
        df = df.drop_duplicates(subset=['username'], keep='first')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cookie.fun Users"
    
    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
    
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
    
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 25
    
    ws.freeze_panes = "A2"
    wb.save(filename)
    print(f"\n✅ Exported {len(df)} users to {filename}")


def main():
    print("=" * 60)
    print("Cookie.fun User Scraper (DrissionPage)")
    print("=" * 60)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    page = create_browser()
    all_data = []
    
    try:
        # Navigate to agents page
        print("📊 Navigating to cookie.fun/agents...")
        page.get("https://www.cookie.fun/agents")
        
        if wait_for_cloudflare(page):
            # Take screenshot
            page.get_screenshot(path="cookie_page.png")
            print("📸 Screenshot saved")
            
            # Wait for page to fully load
            time.sleep(5)
            
            # Scroll to load more content
            print("📜 Scrolling...")
            for i in range(10):
                page.scroll.to_bottom()
                time.sleep(1)
            
            # Extract data
            agents = extract_agents_data(page)
            all_data.extend(agents)
            print(f"  Collected {len(agents)} agents")
        else:
            print("❌ Could not bypass Cloudflare")
            page.get_screenshot(path="cookie_blocked.png")
            
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
    finally:
        page.quit()
    
    # Export
    print("\n" + "=" * 60)
    print("EXPORT")
    print("=" * 60)
    
    if all_data:
        export_to_excel(all_data)
        print(f"\n📈 Total users: {len(all_data)}")
    else:
        print("\n⚠️ No data collected")
        print("The site may be blocking automated access.")
    
    print(f"\nCompleted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
