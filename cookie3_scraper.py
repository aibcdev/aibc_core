#!/usr/bin/env python3
"""
Cookie3/Cookie.fun User Data Scraper

Scrapes user data from Cookie.fun platform including:
- Username (Twitter/X handle)
- User ID
- Onchain wallet address
- Score/Points (Snaps, mindshare)

Usage:
    python cookie3_scraper.py [--api-key YOUR_API_KEY]
"""

import argparse
import json
import time
from datetime import datetime
from typing import Optional

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class Cookie3Scraper:
    """Scraper for Cookie.fun platform data."""
    
    BASE_URL = "https://api.cookie.fun"
    API_VERSION = "v2"
    
    def __init__(self, api_key: Optional[str] = None):
        """Initialize scraper with optional API key."""
        self.api_key = api_key
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            "Accept": "application/json",
            "Content-Type": "application/json",
        })
        if api_key:
            self.session.headers["Authorization"] = f"Bearer {api_key}"
    
    def _make_request(self, endpoint: str, params: dict = None) -> dict:
        """Make API request with error handling."""
        url = f"{self.BASE_URL}/{self.API_VERSION}/{endpoint}"
        try:
            response = self.session.get(url, params=params, timeout=30)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.HTTPError as e:
            print(f"HTTP Error for {endpoint}: {e}")
            return {}
        except requests.exceptions.RequestException as e:
            print(f"Request failed for {endpoint}: {e}")
            return {}
        except json.JSONDecodeError:
            print(f"Invalid JSON response from {endpoint}")
            return {}
    
    def get_agents(self, page: int = 1, limit: int = 100) -> list:
        """Get AI agents with mindshare scores."""
        print(f"Fetching agents page {page}...")
        params = {"page": page, "pageSize": limit, "interval": "_7Days"}
        data = self._make_request("agents/agentsPaged", params)
        return data.get("ok", {}).get("data", []) if data else []
    
    def get_all_agents(self, max_pages: int = 50) -> list:
        """Fetch all agents with pagination."""
        all_agents = []
        page = 1
        
        while page <= max_pages:
            agents = self.get_agents(page=page)
            if not agents:
                break
            all_agents.extend(agents)
            print(f"  Retrieved {len(agents)} agents (total: {len(all_agents)})")
            page += 1
            time.sleep(0.5)  # Rate limiting
        
        return all_agents
    
    def get_creators(self, page: int = 1, limit: int = 100) -> list:
        """Get creators/KOLs with Snaps points."""
        print(f"Fetching creators page {page}...")
        params = {"page": page, "limit": limit}
        data = self._make_request("creators", params)
        return data.get("data", []) if data else []
    
    def get_influencers(self, page: int = 1, limit: int = 100) -> list:
        """Get influencers with engagement scores."""
        print(f"Fetching influencers page {page}...")
        params = {"page": page, "limit": limit}
        data = self._make_request("influencers", params)
        return data.get("data", []) if data else []
    
    def get_leaderboard(self) -> list:
        """Get leaderboard data."""
        print("Fetching leaderboard...")
        data = self._make_request("leaderboard")
        return data.get("data", []) if data else []


def normalize_agent_data(agents: list) -> list:
    """Normalize agent data to standard format."""
    normalized = []
    for agent in agents:
        normalized.append({
            "username": agent.get("twitterUsername", agent.get("name", "")),
            "user_id": agent.get("agentId", agent.get("id", "")),
            "display_name": agent.get("name", ""),
            "onchain_address": agent.get("contracts", [{}])[0].get("contractAddress", "") if agent.get("contracts") else "",
            "mindshare_score": agent.get("mindshare", 0),
            "followers": agent.get("followersCount", 0),
            "market_cap": agent.get("marketCap", 0),
            "volume_24h": agent.get("volume24Hours", 0),
            "category": agent.get("category", "agent"),
            "source": "cookie.fun/agents",
        })
    return normalized


def normalize_creator_data(creators: list) -> list:
    """Normalize creator data to standard format."""
    normalized = []
    for creator in creators:
        normalized.append({
            "username": creator.get("twitterHandle", creator.get("username", "")),
            "user_id": creator.get("userId", creator.get("id", "")),
            "display_name": creator.get("displayName", creator.get("name", "")),
            "onchain_address": creator.get("walletAddress", creator.get("address", "")),
            "snaps_points": creator.get("snapsPoints", creator.get("points", 0)),
            "engagement_score": creator.get("engagementScore", 0),
            "followers": creator.get("followers", 0),
            "rank": creator.get("rank", 0),
            "category": "creator",
            "source": "cookie.fun/creators",
        })
    return normalized


def export_to_excel(data: list, filename: str = "cookie3_users.xlsx"):
    """Export data to formatted Excel file."""
    if not data:
        print("No data to export!")
        return
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Reorder columns for clarity
    column_order = [
        "username", "display_name", "user_id", "onchain_address",
        "mindshare_score", "snaps_points", "engagement_score",
        "followers", "market_cap", "volume_24h", "rank", "category", "source"
    ]
    # Only include columns that exist
    ordered_cols = [c for c in column_order if c in df.columns]
    remaining_cols = [c for c in df.columns if c not in ordered_cols]
    df = df[ordered_cols + remaining_cols]
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cookie3 Users"
    
    # Header styling
    header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Write headers
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Write data
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left")
    
    # Auto-adjust column widths
    for col_idx, column in enumerate(df.columns, 1):
        max_length = max(
            len(str(column)),
            max(len(str(val)) for val in df[column].astype(str)) if len(df) > 0 else 0
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save
    wb.save(filename)
    print(f"\n✅ Exported {len(data)} records to {filename}")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(description="Scrape Cookie3/Cookie.fun user data")
    parser.add_argument("--api-key", help="Cookie.fun API key for authenticated access")
    parser.add_argument("--output", default="cookie3_users.xlsx", help="Output Excel filename")
    parser.add_argument("--max-pages", type=int, default=50, help="Maximum pages to fetch per endpoint")
    args = parser.parse_args()
    
    print("=" * 60)
    print("Cookie3/Cookie.fun User Data Scraper")
    print("=" * 60)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"API Key: {'Provided' if args.api_key else 'Not provided (using public access)'}")
    print()
    
    scraper = Cookie3Scraper(api_key=args.api_key)
    all_data = []
    
    # Fetch agents
    print("\n📊 Fetching AI Agents...")
    agents = scraper.get_all_agents(max_pages=args.max_pages)
    if agents:
        normalized_agents = normalize_agent_data(agents)
        all_data.extend(normalized_agents)
        print(f"  ✓ Found {len(normalized_agents)} agents")
    else:
        print("  ⚠ No agents found (API key may be required)")
    
    # Fetch creators
    print("\n👥 Fetching Creators/KOLs...")
    creators = scraper.get_creators()
    if creators:
        normalized_creators = normalize_creator_data(creators)
        all_data.extend(normalized_creators)
        print(f"  ✓ Found {len(normalized_creators)} creators")
    else:
        print("  ⚠ No creators found (API key may be required)")
    
    # Fetch leaderboard
    print("\n🏆 Fetching Leaderboard...")
    leaderboard = scraper.get_leaderboard()
    if leaderboard:
        # Normalize leaderboard entries
        for entry in leaderboard:
            all_data.append({
                "username": entry.get("username", entry.get("twitterHandle", "")),
                "user_id": entry.get("userId", entry.get("id", "")),
                "display_name": entry.get("displayName", ""),
                "onchain_address": entry.get("walletAddress", ""),
                "snaps_points": entry.get("points", 0),
                "rank": entry.get("rank", 0),
                "category": "leaderboard",
                "source": "cookie.fun/leaderboard",
            })
        print(f"  ✓ Found {len(leaderboard)} leaderboard entries")
    else:
        print("  ⚠ No leaderboard data found")
    
    # Export results
    print("\n" + "=" * 60)
    print("EXPORT")
    print("=" * 60)
    
    if all_data:
        # Remove duplicates based on username
        df = pd.DataFrame(all_data)
        df = df.drop_duplicates(subset=["username"], keep="first")
        unique_data = df.to_dict("records")
        
        export_to_excel(unique_data, args.output)
        
        print(f"\n📈 Summary:")
        print(f"   Total unique users: {len(unique_data)}")
        print(f"   Output file: {args.output}")
    else:
        print("\n❌ No data collected. Please provide an API key or check connectivity.")
        print("\nTo get an API key:")
        print("  1. Sign up at https://cookie.fun")
        print("  2. Request API access from their platform")
        print("  3. Run: python cookie3_scraper.py --api-key YOUR_KEY")
    
    print(f"\nCompleted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
